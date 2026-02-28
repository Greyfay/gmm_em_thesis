#!/usr/bin/env python3
"""Benchmark comparing PyTorch (unoptimized) vs scikit-learn GaussianMixture.

This script compares the runtime of the loop-based PyTorch implementation
against scikit-learn's reference GaussianMixture to show that both produce
equivalent results but with different performance characteristics.
"""

import sys
import os
import time
from typing import Callable, Dict, List, Tuple

import numpy as np
import pandas as pd
import torch
from sklearn.mixture import GaussianMixture

try:
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Add parent directory to path
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, ROOT)

# Import unoptimized PyTorch implementation
from implementation import _v0_ref as torch_impl


def timer(func: Callable, *args, n_runs: int = 5, warmup: int = 1, **kwargs) -> Tuple[float, float]:
    """Time a function with warmup runs.
    
    Returns:
        (mean_time, std_time) in milliseconds
    """
    # Warmup runs
    for _ in range(warmup):
        _ = func(*args, **kwargs)
    
    # Timed runs
    times = []
    for _ in range(n_runs):
        if torch.cuda.is_available():
            torch.cuda.synchronize()
        
        start = time.perf_counter()
        result = func(*args, **kwargs)
        
        if torch.cuda.is_available():
            torch.cuda.synchronize()
        
        end = time.perf_counter()
        times.append((end - start) * 1000)  # Convert to ms
    
    return np.mean(times), np.std(times)


def generate_test_data(
    N: int = 1000,
    D: int = 50,
    K: int = 5,
    seed: int = None,
) -> Tuple[np.ndarray, torch.Tensor]:
    """Generate synthetic test data."""
    if seed is None:
        seed = np.random.randint(1, 1001)
    
    np.random.seed(seed)
    torch.manual_seed(seed)
    
    # Generate data
    X_np = np.random.randn(N, D).astype(np.float64)
    X_torch = torch.from_numpy(X_np).to(torch.float64)
    
    return X_np, X_torch


def benchmark_fit():
    """Benchmark GMM fit for various covariance types and problem sizes."""
    print("\n" + "="*100)
    print("BENCHMARK: TorchGaussianMixture (unoptimized) vs scikit-learn GaussianMixture")
    print("="*100)
    
    results = []
    n_seeds = 5  # Number of random seeds to average over
    
    for cov_type in ["spherical", "diag", "tied", "full"]:
        print(f"\n--- Covariance type: {cov_type} ---")
        
        for N, D, K in [(500, 20, 5), (1000, 50, 5), (2000, 100, 5)]:
            print(f"N={N}, D={D}, K={K}:")
            
            # Collect metrics across seeds
            sklearn_times = []
            sklearn_n_iters = []
            sklearn_converged = []
            sklearn_lower_bounds = []
            sklearn_log_likelihoods = []
            
            torch_times = []
            torch_n_iters = []
            torch_converged = []
            torch_lower_bounds = []
            torch_log_likelihoods = []
            
            sklearn_times_per_iter = []
            torch_times_per_iter = []
            
            for seed_idx in range(n_seeds):
                seed = np.random.randint(1, 10001)
                X_np, X_torch = generate_test_data(N, D, K, seed=seed)
                
                # sklearn fit
                start = time.perf_counter()
                sklearn_model = GaussianMixture(
                    n_components=K,
                    covariance_type=cov_type,
                    max_iter=150,
                    n_init=1,
                    init_params="kmeans",
                    random_state=seed,
                    verbose=0,
                    tol=1e-3,
                )
                sklearn_model.fit(X_np)
                sklearn_time = (time.perf_counter() - start) * 1000  # ms
                
                sklearn_times.append(sklearn_time)
                sklearn_n_iters.append(sklearn_model.n_iter_)
                sklearn_converged.append(sklearn_model.converged_)
                sklearn_lower_bounds.append(sklearn_model.lower_bound_)
                sklearn_log_likelihoods.append(sklearn_model.score(X_np))
                sklearn_times_per_iter.append(sklearn_time / sklearn_model.n_iter_)
                
                # PyTorch fit
                if torch.cuda.is_available():
                    torch.cuda.synchronize()
                start = time.perf_counter()
                torch_model = torch_impl.TorchGaussianMixture(
                    n_components=K,
                    covariance_type=cov_type,
                    max_iter=150,
                    n_init=1,
                    init_params="kmeans",
                    dtype=torch.float64,
                    tol=1e-3,
                )
                torch_model.fit(X_torch)
                if torch.cuda.is_available():
                    torch.cuda.synchronize()
                torch_time = (time.perf_counter() - start) * 1000  # ms
                
                torch_times.append(torch_time)
                torch_n_iters.append(torch_model.n_iter_)
                torch_converged.append(torch_model.converged_)
                torch_lower_bounds.append(torch_model.lower_bound_)
                torch_log_likelihoods.append(float(torch_model.score(X_torch).item()))
                torch_times_per_iter.append(torch_time / torch_model.n_iter_)
            
            # Compute averages
            sklearn_time_mean = np.mean(sklearn_times)
            sklearn_time_std = np.std(sklearn_times)
            sklearn_n_iter_mean = np.mean(sklearn_n_iters)
            sklearn_converged_rate = np.mean(sklearn_converged)
            sklearn_lower_bound_mean = np.mean(sklearn_lower_bounds)
            sklearn_log_likelihood_mean = np.mean(sklearn_log_likelihoods)
            sklearn_time_per_iter_mean = np.mean(sklearn_times_per_iter)
            sklearn_time_per_iter_std = np.std(sklearn_times_per_iter)
            
            torch_time_mean = np.mean(torch_times)
            torch_time_std = np.std(torch_times)
            torch_n_iter_mean = np.mean(torch_n_iters)
            torch_converged_rate = np.mean(torch_converged)
            torch_lower_bound_mean = np.mean(torch_lower_bounds)
            torch_log_likelihood_mean = np.mean(torch_log_likelihoods)
            torch_time_per_iter_mean = np.mean(torch_times_per_iter)
            torch_time_per_iter_std = np.std(torch_times_per_iter)
            
            speedup = sklearn_time_mean / torch_time_mean
            
            print(f"  sklearn: {sklearn_time_mean:.1f}±{sklearn_time_std:.1f}ms, "
                  f"{sklearn_n_iter_mean:.1f} iters, {sklearn_time_per_iter_mean:.2f}±{sklearn_time_per_iter_std:.2f}ms/iter, "
                  f"converged: {sklearn_converged_rate*100:.0f}%")
            print(f"  pytorch: {torch_time_mean:.1f}±{torch_time_std:.1f}ms, "
                  f"{torch_n_iter_mean:.1f} iters, {torch_time_per_iter_mean:.2f}±{torch_time_per_iter_std:.2f}ms/iter, "
                  f"converged: {torch_converged_rate*100:.0f}%")
            print(f"  Ratio: {speedup:.2f}x")
            
            results.append({
                "Covariance Type": cov_type,
                "N": N,
                "D": D,
                "K": K,
                "sklearn Time (ms)": sklearn_time_mean,
                "sklearn Time Std (ms)": sklearn_time_std,
                "sklearn n_iter": sklearn_n_iter_mean,
                "sklearn Converged (%)": sklearn_converged_rate * 100,
                "sklearn Lower Bound": sklearn_lower_bound_mean,
                "sklearn Log Likelihood": sklearn_log_likelihood_mean,
                "sklearn Time/Iter (ms)": sklearn_time_per_iter_mean,
                "sklearn Time/Iter Std (ms)": sklearn_time_per_iter_std,
                "PyTorch Time (ms)": torch_time_mean,
                "PyTorch Time Std (ms)": torch_time_std,
                "PyTorch n_iter": torch_n_iter_mean,
                "PyTorch Converged (%)": torch_converged_rate * 100,
                "PyTorch Lower Bound": torch_lower_bound_mean,
                "PyTorch Log Likelihood": torch_log_likelihood_mean,
                "PyTorch Time/Iter (ms)": torch_time_per_iter_mean,
                "PyTorch Time/Iter Std (ms)": torch_time_per_iter_std,
                "Ratio (sklearn/pytorch)": speedup,
                "n_seeds": n_seeds,
            })
    
    return results


def benchmark_comprehensive_convergence():
    """Unified benchmark collecting iterations, time, and quality metrics in one pass.
    
    Single pass through all configurations to measure:
    - Iterations to convergence
    - Wall-clock time to convergence
    - Model quality (mean log-likelihood)
    
    Tests across all covariance types using the same configuration matrix as compare_old_vs_new:
    (N, D) = (1k, 20), (1k, 50), (1k, 100), (10k, 20), (10k, 50), (10k, 100),
             (100k, 20), (100k, 50), (100k, 100)
    
    Returns:
        dict with keys 'iterations', 'wall_clock', 'quality' containing respective results
    """
    print("\n" + "="*100)
    print("BENCHMARK: Comprehensive Convergence Analysis (Iterations, Time, Quality)")
    print("="*100)
    
    convergence_results = []
    wall_clock_results = []
    quality_results = []
    
    K = 5  # Fixed number of components
    n_seeds = 3  # Number of random seeds per configuration
    
    # Test configurations: (N, D) as in compare_old_vs_new.py
    test_configs = [
        (1000, 20),      # N=1k, D=20
        (10000, 20),     # N=10k, D=20
        (100000, 20),    # N=100k, D=20
        (1000, 50),      # N=1k, D=50
        (10000, 50),     # N=10k, D=50
        (100000, 50),    # N=100k, D=50
        (1000, 100),     # N=1k, D=100
        (10000, 100),    # N=10k, D=100
        (100000, 100),   # N=100k, D=100
    ]
    
    for N, D in test_configs:
        for cov_type in ["spherical", "diag", "tied", "full"]:
            print(f"\n--- Config: N={N}, D={D}, K={K}, cov_type={cov_type} ---")
            
            # Storage for per-seed metrics
            sklearn_n_iters_list = []
            sklearn_converged_list = []
            sklearn_wall_times = []
            sklearn_log_likelihoods = []
            
            torch_n_iters_list = []
            torch_converged_list = []
            torch_wall_times = []
            torch_log_likelihoods = []
            
            for seed_idx in range(n_seeds):
                seed = np.random.randint(1, 10001)
                X_np, X_torch = generate_test_data(N, D, K, seed=seed)
                
                # sklearn fit - single pass captures everything
                start = time.perf_counter()
                sklearn_model = GaussianMixture(
                    n_components=K,
                    covariance_type=cov_type,
                    max_iter=150,
                    n_init=1,
                    init_params="kmeans",
                    random_state=seed,
                    verbose=0,
                    tol=1e-3,
                )
                sklearn_model.fit(X_np)
                sklearn_wall_time = (time.perf_counter() - start) * 1000  # ms
                
                sklearn_n_iters_list.append(sklearn_model.n_iter_)
                sklearn_converged_list.append(sklearn_model.converged_)
                sklearn_wall_times.append(sklearn_wall_time)
                sklearn_log_likelihoods.append(sklearn_model.score(X_np))
                
                # PyTorch fit - single pass captures everything
                if torch.cuda.is_available():
                    torch.cuda.synchronize()
                start = time.perf_counter()
                torch_model = torch_impl.TorchGaussianMixture(
                    n_components=K,
                    covariance_type=cov_type,
                    max_iter=150,
                    n_init=1,
                    init_params="kmeans",
                    dtype=torch.float64,
                    tol=1e-3,
                )
                torch_model.fit(X_torch)
                if torch.cuda.is_available():
                    torch.cuda.synchronize()
                torch_wall_time = (time.perf_counter() - start) * 1000  # ms
                
                torch_n_iters_list.append(torch_model.n_iter_)
                torch_converged_list.append(torch_model.converged_)
                torch_wall_times.append(torch_wall_time)
                torch_log_likelihoods.append(float(torch_model.score(X_torch).item()))
            
            # Calculate statistics for convergence metrics
            sklearn_n_iter_mean = np.mean(sklearn_n_iters_list)
            sklearn_n_iter_std = np.std(sklearn_n_iters_list)
            sklearn_converged_rate = np.mean(sklearn_converged_list)
            
            torch_n_iter_mean = np.mean(torch_n_iters_list)
            torch_n_iter_std = np.std(torch_n_iters_list)
            torch_converged_rate = np.mean(torch_converged_list)
            
            iter_diff = torch_n_iter_mean - sklearn_n_iter_mean
            iter_diff_pct = (iter_diff / sklearn_n_iter_mean * 100) if sklearn_n_iter_mean > 0 else 0
            
            print(f"  Iterations - sklearn: {sklearn_n_iter_mean:.1f} ± {sklearn_n_iter_std:.1f}, pytorch: {torch_n_iter_mean:.1f} ± {torch_n_iter_std:.1f}")
            
            convergence_results.append({
                "N": N,
                "D": D,
                "K": K,
                "Covariance Type": cov_type,
                "sklearn n_iter": sklearn_n_iter_mean,
                "sklearn n_iter Std": sklearn_n_iter_std,
                "sklearn Converged (%)": sklearn_converged_rate * 100,
                "PyTorch n_iter": torch_n_iter_mean,
                "PyTorch n_iter Std": torch_n_iter_std,
                "PyTorch Converged (%)": torch_converged_rate * 100,
                "Iteration Difference": iter_diff,
                "Iteration Difference (%)": iter_diff_pct,
                "n_seeds": n_seeds,
            })
            
            # Calculate statistics for wall-clock time
            sklearn_wall_time_mean = np.mean(sklearn_wall_times)
            sklearn_wall_time_std = np.std(sklearn_wall_times)
            torch_wall_time_mean = np.mean(torch_wall_times)
            torch_wall_time_std = np.std(torch_wall_times)
            speedup = sklearn_wall_time_mean / torch_wall_time_mean
            
            print(f"  Wall-clock - sklearn: {sklearn_wall_time_mean:.1f}±{sklearn_wall_time_std:.1f}ms, pytorch: {torch_wall_time_mean:.1f}±{torch_wall_time_std:.1f}ms (ratio: {speedup:.2f}x)")
            
            wall_clock_results.append({
                "N": N,
                "D": D,
                "K": K,
                "Covariance Type": cov_type,
                "sklearn Wall-clock Time (ms)": sklearn_wall_time_mean,
                "sklearn Wall-clock Time Std (ms)": sklearn_wall_time_std,
                "PyTorch Wall-clock Time (ms)": torch_wall_time_mean,
                "PyTorch Wall-clock Time Std (ms)": torch_wall_time_std,
                "Ratio (sklearn/pytorch)": speedup,
                "n_seeds": n_seeds,
            })
            
            # Calculate statistics for quality metrics
            sklearn_ll_mean = np.mean(sklearn_log_likelihoods)
            sklearn_ll_std = np.std(sklearn_log_likelihoods)
            torch_ll_mean = np.mean(torch_log_likelihoods)
            torch_ll_std = np.std(torch_log_likelihoods)
            ll_diff = torch_ll_mean - sklearn_ll_mean
            ll_diff_pct = (abs(ll_diff) / abs(sklearn_ll_mean) * 100) if sklearn_ll_mean != 0 else 0
            
            print(f"  Quality - sklearn: {sklearn_ll_mean:.4f}±{sklearn_ll_std:.4f}, pytorch: {torch_ll_mean:.4f}±{torch_ll_std:.4f} (diff: {ll_diff:+.4f})")
            
            quality_results.append({
                "N": N,
                "D": D,
                "K": K,
                "Covariance Type": cov_type,
                "sklearn Mean Log-Likelihood": sklearn_ll_mean,
                "sklearn Mean Log-Likelihood Std": sklearn_ll_std,
                "sklearn n_iter": sklearn_n_iter_mean,
                "sklearn Converged (%)": sklearn_converged_rate * 100,
                "PyTorch Mean Log-Likelihood": torch_ll_mean,
                "PyTorch Mean Log-Likelihood Std": torch_ll_std,
                "PyTorch n_iter": torch_n_iter_mean,
                "PyTorch Converged (%)": torch_converged_rate * 100,
                "Log-Likelihood Difference": ll_diff,
                "Log-Likelihood Difference (%)": ll_diff_pct,
                "n_seeds": n_seeds,
            })
    
    return {
        "iterations": convergence_results,
        "wall_clock": wall_clock_results,
        "quality": quality_results,
    }


def benchmark_individual_functions():
    """Benchmark individual functions where applicable."""
    print("\n" + "="*100)
    print("BENCHMARK: Individual function comparisons")
    print("="*100)
    
    results = []
    
    # Test data
    N, D, K = 1000, 50, 5
    random_seed = np.random.randint(1, 1001)
    
    print(f"\nUsing N={N}, D={D}, K={K}, seed={random_seed}")
    
    # Benchmark E-step (log probability computation)
    print(f"\n--- E-step: Log probability computation ---")
    
    for cov_type in ["spherical", "diag", "tied", "full"]:
        np.random.seed(random_seed)
        torch.manual_seed(random_seed)
        
        X_np = np.random.randn(N, D).astype(np.float64)
        X_torch = torch.from_numpy(X_np).to(torch.float64)
        
        means_np = np.random.randn(K, D).astype(np.float64)
        means_torch = torch.from_numpy(means_np).to(torch.float64)
        
        weights_np = np.ones(K) / K
        weights_torch = torch.from_numpy(weights_np).to(torch.float64)
        
        # Generate covariances
        if cov_type == "spherical":
            cov_torch = torch.ones((K,), dtype=torch.float64) + 0.5
        elif cov_type == "diag":
            cov_torch = torch.ones((K, D), dtype=torch.float64) + 0.5
        elif cov_type == "tied":
            cov_torch = torch.eye(D, dtype=torch.float64) + 0.1
        else:  # full
            cov_torch = torch.stack([
                torch.eye(D, dtype=torch.float64) + 0.1 * torch.randn(D, D)
                for _ in range(K)
            ])
            cov_torch = torch.bmm(cov_torch, cov_torch.transpose(-1, -2))
        
        # PyTorch E-step
        def torch_estep():
            lower, log_resp = torch_impl._expectation_step_precchol(
                X_torch, means_torch, 
                torch_impl._compute_precisions_cholesky(cov_torch, cov_type),
                weights_torch, cov_type
            )
            return lower, log_resp
        
        torch_time, torch_std = timer(torch_estep, n_runs=5, warmup=2)
        
        print(f"{cov_type:9s}: {torch_time:.3f} ± {torch_std:.3f} ms")
        
        results.append({
            "Function": "E-step",
            "Covariance Type": cov_type,
            "N": N,
            "D": D,
            "K": K,
            "PyTorch Time (ms)": torch_time,
            "PyTorch Std (ms)": torch_std,
        })
    
    return results


def main():
    """Run all benchmarks."""
    print("="*100)
    print("PYTORCH (UNOPTIMIZED) vs SCIKIT-LEARN COMPARISON")
    print("="*100)
    print(f"PyTorch version: {torch.__version__}")
    print(f"NumPy version: {np.__version__}")
    print(f"scikit-learn version: (imported)")
    print(f"Device: {torch.device('cuda' if torch.cuda.is_available() else 'cpu')}")
    
    # Run benchmarks
    fit_results = benchmark_fit()
    comprehensive_results = benchmark_comprehensive_convergence()
    func_results = benchmark_individual_functions()
    
    # Extract individual result types from comprehensive benchmark
    convergence_results = comprehensive_results["iterations"]
    wall_clock_results = comprehensive_results["wall_clock"]
    quality_results = comprehensive_results["quality"]
    
    print("\n" + "="*100)
    print("BENCHMARKS COMPLETE")
    print("="*100)
    
    # Create DataFrames
    df_fit = pd.DataFrame(fit_results)
    
    # Save to Excel with enhanced formatting
    output_file = os.path.join(os.path.dirname(__file__), "pytorch_vs_sklearn.xlsx")
    
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        # Write fit results
        df_fit.to_excel(writer, sheet_name="Full Fit Comparison", index=False)
        
        # Write convergence results
        if convergence_results:
            df_convergence = pd.DataFrame(convergence_results)
            # Sort by covariance type, then N, then D
            cov_type_order = {"spherical": 0, "diag": 1, "tied": 2, "full": 3}
            df_convergence["_cov_order"] = df_convergence["Covariance Type"].map(cov_type_order)
            df_convergence = df_convergence.sort_values(["_cov_order", "N", "D"]).drop("_cov_order", axis=1)
            df_convergence.to_excel(writer, sheet_name="Convergence Iterations", index=False)
        
        # Write function results if available
        if func_results:
            df_func = pd.DataFrame(func_results)
            df_func.to_excel(writer, sheet_name="Function Comparison", index=False)
        
        # Write wall-clock time to convergence results
        if wall_clock_results:
            df_wall_clock = pd.DataFrame(wall_clock_results)
            # Sort by covariance type, then N, then D
            cov_type_order = {"spherical": 0, "diag": 1, "tied": 2, "full": 3}
            df_wall_clock["_cov_order"] = df_wall_clock["Covariance Type"].map(cov_type_order)
            df_wall_clock = df_wall_clock.sort_values(["_cov_order", "N", "D"]).drop("_cov_order", axis=1)
            df_wall_clock.to_excel(writer, sheet_name="Wall-clock Time to Convergence", index=False)
        
        # Write quality at convergence results
        if quality_results:
            df_quality = pd.DataFrame(quality_results)
            # Sort by covariance type, then N, then D
            cov_type_order = {"spherical": 0, "diag": 1, "tied": 2, "full": 3}
            df_quality["_cov_order"] = df_quality["Covariance Type"].map(cov_type_order)
            df_quality = df_quality.sort_values(["_cov_order", "N", "D"]).drop("_cov_order", axis=1)
            df_quality.to_excel(writer, sheet_name="Quality at Convergence", index=False)
        
        # Create summary statistics sheet
        summary_data = {
            "Metric": [
                "Total Benchmarks",
                "Seeds per Configuration",
                "Average Ratio (sklearn/pytorch)",
                "Max Ratio (sklearn faster)",
                "Min Ratio (pytorch faster)",
                "Avg sklearn Time/Iter (ms)",
                "Avg PyTorch Time/Iter (ms)",
                "Avg sklearn n_iter",
                "Avg PyTorch n_iter",
                "Avg sklearn Converged (%)",
                "Avg PyTorch Converged (%)",
            ],
            "Value": [
                len(df_fit),
                df_fit['n_seeds'].iloc[0],
                f"{df_fit['Ratio (sklearn/pytorch)'].mean():.2f}x",
                f"{df_fit['Ratio (sklearn/pytorch)'].max():.2f}x",
                f"{df_fit['Ratio (sklearn/pytorch)'].min():.2f}x",
                f"{df_fit['sklearn Time/Iter (ms)'].mean():.3f}",
                f"{df_fit['PyTorch Time/Iter (ms)'].mean():.3f}",
                f"{df_fit['sklearn n_iter'].mean():.1f}",
                f"{df_fit['PyTorch n_iter'].mean():.1f}",
                f"{df_fit['sklearn Converged (%)'].mean():.1f}%",
                f"{df_fit['PyTorch Converged (%)'].mean():.1f}%",
            ]
        }
        df_summary = pd.DataFrame(summary_data)
        df_summary.to_excel(writer, sheet_name="Summary", index=False)
        
        # Create breakdown by covariance type
        breakdown_data = []
        for cov_type in df_fit["Covariance Type"].unique():
            subset = df_fit[df_fit["Covariance Type"] == cov_type]
            breakdown_data.append({
                "Covariance Type": cov_type,
                "Num Benchmarks": len(subset),
                "Avg sklearn Time (ms)": subset["sklearn Time (ms)"].mean(),
                "Avg PyTorch Time (ms)": subset["PyTorch Time (ms)"].mean(),
                "Avg sklearn Time/Iter (ms)": subset["sklearn Time/Iter (ms)"].mean(),
                "Avg PyTorch Time/Iter (ms)": subset["PyTorch Time/Iter (ms)"].mean(),
                "Avg sklearn n_iter": subset["sklearn n_iter"].mean(),
                "Avg PyTorch n_iter": subset["PyTorch n_iter"].mean(),
                "Avg Ratio (sklearn/pytorch)": subset["Ratio (sklearn/pytorch)"].mean(),
            })
        df_breakdown = pd.DataFrame(breakdown_data)
        df_breakdown.to_excel(writer, sheet_name="By Covariance Type", index=False)
        
        # Apply formatting if openpyxl available
        if OPENPYXL_AVAILABLE:
            _format_excel_sheets(writer, df_fit, df_func if func_results else None)
    
    print(f"\n✓ Results exported to: {output_file}")
    print(f"\nFull Fit Summary:")
    print(f"  Total benchmarks: {len(df_fit)}")
    print(f"  Seeds per config: {df_fit['n_seeds'].iloc[0]}")
    print(f"  Average ratio (sklearn/pytorch): {df_fit['Ratio (sklearn/pytorch)'].mean():.2f}x")
    print(f"  Max ratio: {df_fit['Ratio (sklearn/pytorch)'].max():.2f}x")
    print(f"  Min ratio: {df_fit['Ratio (sklearn/pytorch)'].min():.2f}x")
    
    # Print interpretation
    ratio_mean = df_fit['Ratio (sklearn/pytorch)'].mean()
    if ratio_mean > 1:
        print(f"\n→ scikit-learn is faster by {ratio_mean:.2f}x on average")
    else:
        print(f"\n→ PyTorch is faster by {1/ratio_mean:.2f}x on average")
    
    # Convergence stats
    print(f"\nConvergence Stats:")
    print(f"  sklearn avg n_iter: {df_fit['sklearn n_iter'].mean():.1f}")
    print(f"  pytorch avg n_iter: {df_fit['PyTorch n_iter'].mean():.1f}")
    print(f"  sklearn converged: {df_fit['sklearn Converged (%)'].mean():.1f}%")
    print(f"  pytorch converged: {df_fit['PyTorch Converged (%)'].mean():.1f}%")
    
    # Time per iteration
    print(f"\nTime per Iteration:")
    print(f"  sklearn: {df_fit['sklearn Time/Iter (ms)'].mean():.3f} ± {df_fit['sklearn Time/Iter Std (ms)'].mean():.3f} ms/iter")
    print(f"  pytorch: {df_fit['PyTorch Time/Iter (ms)'].mean():.3f} ± {df_fit['PyTorch Time/Iter Std (ms)'].mean():.3f} ms/iter")
    
    # Breakdown by covariance type
    print(f"\nBreakdown by Covariance Type:")
    for cov_type in sorted(df_fit["Covariance Type"].unique()):
        subset = df_fit[df_fit["Covariance Type"] == cov_type]
        avg_ratio = subset["Ratio (sklearn/pytorch)"].mean()
        avg_sklearn_time_per_iter = subset["sklearn Time/Iter (ms)"].mean()
        avg_sklearn_time_per_iter_std = subset["sklearn Time/Iter Std (ms)"].mean()
        avg_pytorch_time_per_iter = subset["PyTorch Time/Iter (ms)"].mean()
        avg_pytorch_time_per_iter_std = subset["PyTorch Time/Iter Std (ms)"].mean()
        print(f"  {cov_type:6s}: {avg_ratio:.2f}x ratio, "
              f"sklearn {avg_sklearn_time_per_iter:.3f}±{avg_sklearn_time_per_iter_std:.3f}ms/iter, "
              f"pytorch {avg_pytorch_time_per_iter:.3f}±{avg_pytorch_time_per_iter_std:.3f}ms/iter")


def _format_excel_sheets(writer, df_fit, df_func):
    """Apply formatting to Excel sheets."""
    # Format Full Fit Comparison sheet
    ws = writer.sheets["Full Fit Comparison"]
    
    # Set column widths (adjust for new columns)
    for col_idx in range(1, ws.max_column + 1):
        ws.column_dimensions[chr(64 + col_idx)].width = 18
    
    # Format header row
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Format data cells
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            # Format numeric columns
            if cell.column in [2, 3, 4]:  # N, D, K columns
                cell.number_format = "0"
            elif cell.column in [7, 15]:  # n_iter columns
                cell.number_format = "0.0"
            elif cell.column in [8, 16]:  # Converged % columns
                cell.number_format = "0.0"
            elif cell.column >= 5:  # Other numeric columns
                cell.number_format = "0.000"
    
    # Color code the ratio column (last column)
    ratio_col = ws.max_column - 1  # Ratio is second to last
    for row_num, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, 
                                                min_col=ratio_col, max_col=ratio_col), start=2):
        try:
            ratio = float(row[0].value)
            if ratio > 1.1:
                row[0].fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
            elif ratio < 0.9:
                row[0].fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        except (ValueError, TypeError):
            pass
    
    # Format Summary sheet if it exists
    if "Summary" in writer.sheets:
        ws_summary = writer.sheets["Summary"]
        for col in ["A", "B"]:
            ws_summary.column_dimensions[col].width = 35
        
        for cell in ws_summary[1]:
            cell.fill = header_fill
            cell.font = header_font
    
    # Format By Covariance Type sheet
    if "By Covariance Type" in writer.sheets:
        ws_cov = writer.sheets["By Covariance Type"]
        for col_idx in range(1, ws_cov.max_column + 1):
            ws_cov.column_dimensions[chr(64 + col_idx)].width = 22
        
        for cell in ws_cov[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")


if __name__ == "__main__":
    main()
